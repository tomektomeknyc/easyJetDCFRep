from openpyxl import load_workbook
import csv

# Load the Excel workbook and the DCF sheet
excel_path = "attached_assets/EasyJet- complete.xlsx"
wb = load_workbook(excel_path, data_only=True)
ws = wb["DCF"]  # Ensure this matches your sheet name exactly

# Collect all string-type cell values and their positions
labels_with_pos = []
for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
    for col_idx, cell in enumerate(row):
        if isinstance(cell, str) and cell.strip():  # Check for non-empty strings
            label = cell.strip()
            row_num = row_idx + 1
            col_num = col_idx + 1
            labels_with_pos.append((label, row_num, col_num))
            print(f"'{label}' found at row {row_num}, column {col_num}")

# Write results to a CSV file
csv_path = "label_locations.csv"
with open(csv_path, mode="w", newline="", encoding="utf-8") as f:
    writer = csv.writer(f)
    writer.writerow(["Label", "Row", "Column"])
    writer.writerows(labels_with_pos)

print(f"\n✅ Exported {len(labels_with_pos)} labels to '{csv_path}'")
